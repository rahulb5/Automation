from openpyxl import load_workbook
import csv
from datetime import date, timedelta


def get_date_list(holidays, custom, today = date.today()):
    month = ['NULL' ,'JAN' , 'FEB' , 'MAR' , 'APR' , 'MAY' , 'JUN' , 'JUL' , 'AUG' , 'SEP' , 'OCT' , 'NOV' , 'DEC']
    temp = ""    
    required_days = []
    #D-1    
    working_days = 1
    flag = 0
    while flag == 0:
        day_minus = today - timedelta(days = working_days)
        if day_minus not in holidays and day_minus.weekday() < 5: 
           required_days.append(day_minus)     
           flag = 1
        else: 
            working_days = working_days + 1
    #D-2
    working_days = working_days + 1
    flag = 0
    while flag == 0:
        day_minus = today - timedelta(days = working_days)
        if day_minus not in holidays and day_minus.weekday() < 5: 
           required_days.append(day_minus)     
           flag = 1
        else: 
            working_days = working_days + 1
    
    #D-3
    working_days = working_days + 1
    flag = 0
    while flag == 0:
        day_minus = today - timedelta(days = working_days)
        if day_minus not in holidays and day_minus.weekday() < 5: 
           required_days.append(day_minus)     
           flag = 1
        else: 
            working_days = working_days + 1
    #D-5
    count = 0
    while count < 2: 
        working_days = working_days + 1
        day_minus = today - timedelta(days = working_days)
        if day_minus not in holidays and day_minus.weekday() < 5:
            count += 1
        if count == 2:
            required_days.append(day_minus)
    
    #D-10
    count = 0
    while count < 5: 
        working_days = working_days + 1
        day_minus = today - timedelta(days = working_days)
        if day_minus not in holidays and day_minus.weekday() < 5:
            count += 1
        if count == 5:
            required_days.append(day_minus)
    
    if custom not in holidays and custom.weekday() < 5:
        required_days.append(custom)
    
    src_list = []
    temp = ""
    for i in required_days:    
        if i.day < 10:
            temp = "D:/MATA/India/Data/Bhavcopy/NSE-EOD/cm0" + str(i.day) + month[i.month] + str(i.year) + "bhav.csv"
        else:
            temp = "D:/MATA/India/Data/Bhavcopy/NSE-EOD/cm" + str(i.day) + month[i.month] + str(i.year) + "bhav.csv"
        
        src_list.append(temp)
    
    return src_list
    
dest= "E:/Data/Nifty FNO Scrips_automate.xlsm"

wb = load_workbook(dest, keep_vba=True, keep_links= False, data_only=True)
ws = wb["FNO_Historic_Data"]
ws_working = wb["Working"]

count = 3 
holidays = []

while ws_working.cell(count, 2).value != None:
    holidays.append(ws_working.cell(count,2).value)
    count = count + 1

custom = ws.cell(1,7).value
src_list = get_date_list(holidays , custom, ws.cell(1,1).value)

companies = []
for i in range(3,ws.max_row+1) :
    companies.append(ws.cell(i,1).value)
    
for data_source in range(0,len(src_list)):
      
    d = open(src_list[data_source])
    reader = csv.reader(d)

    data = {}

    for value in reader:
        if value[1] == 'EQ':
            data[value[0]] = float(value[5])
        
    data_from = []
    for company in companies:
        if company in data.keys() :
            data_from.append(data[company])

    for i in range(3, ws.max_row+1):
        ws.cell(i,data_source + 2).value = data_from[i-3]
    

wb.save(dest)
