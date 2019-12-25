from openpyxl import load_workbook
import csv

dest = "E:/Data/Mutual Fund Master automate.xlsm"

wb = load_workbook(dest,keep_vba=True, keep_links= False, data_only=True)
ws = wb["NAV_Historic_Data"]

count = 2
required_dates = []
while ws.cell(2 , count).value != None:
    required_dates.append(ws.cell(2,count).value)
    count += 1

required_codes = []
for i in range(3, ws.max_row):
    required_codes.append(ws.cell(i , 1).value)

#count = 0
for extract_date in required_dates :
    temp = []
    if extract_date.month < 10:
        temp.append("0" + str(extract_date.month))
    else:
        temp.append(str(extract_date.month))
    if extract_date.day < 10:
        temp.append("0" + str(extract_date.day))
    else:
        temp.append(str(extract_date.day))
    src = "E:/Data/NAV History Report/" + str(extract_date.year) + temp[0] + temp[1] + "_NAVHistoryReport.csv"
    try:
        d = open(src)
    except:
        continue
    
    reader = csv.reader(d)
    
    for j in reader:
        try:
            key = int(j[0])
            index = required_codes.index(key) + 3
            col_index = required_dates.index(extract_date) + 2
            ws.cell(index , col_index).value = float(j[4])
        except:
            continue
    
wb.save(dest)